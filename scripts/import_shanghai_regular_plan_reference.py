from __future__ import annotations

import json
import re
from collections import Counter
from datetime import date
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


REPO_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_INPUT = Path(
    "/Users/tl/Library/Containers/com.tencent.xinWeChat/Data/Documents/xwechat_files/"
    "wxid_icnw0n1ekpvz21_30fa/temp/drag/2026普通批整理版.xlsx"
)
OUTPUT_PATH = REPO_ROOT / "data/shanghai/regular-2026-plan-reference.json"
OFFICIAL_ADMISSIONS_PATH = REPO_ROOT / "data/shanghai/all-admissions.json"

SOURCE_LABEL = "用户提供《2026普通批整理版.xlsx》"
SOURCE_QUALITY = "用户提供整理版，按原表字段结构化；作为第三方参考，需回当年招生专业目录和考试院公告核对。"


def clean(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if text.endswith(".0") and text[:-2].isdigit():
        return text[:-2]
    return re.sub(r"\s+", " ", text)


def parse_int(value: Any) -> int | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(value) if value.is_integer() else None
    text = clean(value).replace(",", "")
    if not text or text in {"-", "—", "待定", "未公布"}:
        return None
    match = re.search(r"\d+", text)
    return int(match.group(0)) if match else None


def parse_score_label(value: Any) -> tuple[str, int | None, str | None]:
    label = clean(value)
    if not label:
        return "", None, None
    score = parse_int(label)
    if score is None:
        return label, None, None
    score_type = "threshold" if "以上" in label or "及以上" in label else "exact"
    return label, score, score_type


def get_group_suffix(group_name: str) -> str:
    match = re.search(r"\(([^)]+)\)|（([^）]+)）", group_name)
    return match.group(1) or match.group(2) if match else ""


def split_major_name(major_name: str) -> tuple[str, str]:
    match = re.match(r"^(\d{1,3})[-－](.+)$", major_name)
    if not match:
        return "", major_name
    return match.group(1), clean(match.group(2))


def load_official_2026_groups() -> dict[str, dict[str, Any]]:
    dataset = json.loads(OFFICIAL_ADMISSIONS_PATH.read_text(encoding="utf-8"))
    by_group_name: dict[str, dict[str, Any]] = {}
    duplicates: set[str] = set()
    for record in dataset["records"]:
        if record.get("year") != 2026:
            continue
        group_name = clean(record.get("groupName"))
        if not group_name:
            continue
        if group_name in by_group_name:
            duplicates.add(group_name)
        by_group_name[group_name] = record
    for duplicate in duplicates:
        by_group_name.pop(duplicate, None)
    return by_group_name


def build_dataset(input_path: Path) -> dict[str, Any]:
    official_groups = load_official_2026_groups()
    workbook = load_workbook(input_path, read_only=True, data_only=True)
    sheet = workbook.active

    headers = [clean(cell.value) for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    index = {header: offset for offset, header in enumerate(headers)}
    required = ["省份", "类别", "院校名称", "院校专业组", "选科要求", "专业名称", "26计划", "26投档", "学费", "25录取", "25最低", "25平均", "25位次"]
    missing = [header for header in required if header not in index]
    if missing:
        raise ValueError(f"missing required headers: {', '.join(missing)}")

    groups: list[dict[str, Any]] = []
    unmatched_groups: list[dict[str, Any]] = []
    current_group: dict[str, Any] | None = None

    for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        school_region = clean(row[index["省份"]])
        school_name = clean(row[index["院校名称"]])
        group_name = clean(row[index["院校专业组"]])
        major_name = clean(row[index["专业名称"]])

        if not major_name:
            if not school_name or not group_name:
                current_group = None
                continue

            official = official_groups.get(group_name)
            line_label, line_score, line_type = parse_score_label(row[index["26投档"]])
            current_group = {
                "schoolSlug": clean(official.get("schoolSlug")) if official else "",
                "schoolName": school_name,
                "schoolAlias": clean(official.get("schoolAlias")) if official else school_name,
                "schoolRegion": school_region,
                "category": clean(row[index["类别"]]),
                "groupCode": clean(official.get("groupCode")) if official else "",
                "groupName": group_name,
                "groupSuffix": get_group_suffix(group_name),
                "subjectRequirement": clean(row[index["选科要求"]]),
                "groupPlan2026": parse_int(row[index["26计划"]]),
                "groupLine2026Label": line_label,
                "groupLine2026": line_score,
                "groupLine2026Type": line_type,
                "groupAdmittedCount2025": parse_int(row[index["25录取"]]),
                "groupMinScore2025Label": clean(row[index["25最低"]]),
                "groupMinScore2025": parse_int(row[index["25最低"]]),
                "groupAverageScore2025": parse_int(row[index["25平均"]]),
                "groupAverageRank2025": parse_int(row[index["25位次"]]),
                "officialSourceType": clean(official.get("sourceType")) if official else "",
                "officialSourceUrl": clean(official.get("sourceUrl")) if official else "",
                "majors": [],
            }
            groups.append(current_group)
            if not official:
                unmatched_groups.append(
                    {
                        "rowNumber": row_number,
                        "schoolName": school_name,
                        "groupName": group_name,
                        "subjectRequirement": current_group["subjectRequirement"],
                        "groupPlan2026": current_group["groupPlan2026"],
                    }
                )
            continue

        if current_group is None:
            continue

        admitted_count = parse_int(row[index["25录取"]])
        average_score = parse_int(row[index["25平均"]])
        average_rank = parse_int(row[index["25位次"]])
        major_code, display_major_name = split_major_name(major_name)
        current_group["majors"].append(
            {
                "sourceRow": row_number,
                "majorCode": major_code,
                "majorName": display_major_name,
                "plan2026": parse_int(row[index["26计划"]]),
                "tuition": parse_int(row[index["学费"]]),
                "admittedCount2025": admitted_count,
                "minScoreLabel": clean(row[index["25最低"]]),
                "minScore2025": parse_int(row[index["25最低"]]),
                "minRankLabel": clean(row[index["25位次"]]),
                "averageScore2025": average_score,
                "averageRank2025": average_rank,
            }
        )

    workbook.close()
    category_counts = Counter(group["category"] for group in groups if group["category"])
    subject_counts = Counter(group["subjectRequirement"] for group in groups if group["subjectRequirement"])
    matched_count = len(groups) - len(unmatched_groups)
    major_count = sum(len(group["majors"]) for group in groups)

    return {
        "meta": {
            "region": "上海",
            "year": 2026,
            "referenceAdmissionYear": 2025,
            "batch": "本科普通批次",
            "grain": "院校专业组+专业计划",
            "sourceTrust": "third-party-reference",
            "sourceLabel": SOURCE_LABEL,
            "sourceFile": input_path.name,
            "sourceQuality": SOURCE_QUALITY,
            "generatedAt": date.today().isoformat(),
            "recordCount": major_count,
            "groupCount": len(groups),
            "schoolCount": len({group["schoolName"] for group in groups}),
            "matchedOfficialGroupCount": matched_count,
            "unmatchedGroupCount": len(unmatched_groups),
            "groupPlan2026Total": sum(group["groupPlan2026"] or 0 for group in groups),
            "majorPlan2026Total": sum((major["plan2026"] or 0) for group in groups for major in group["majors"]),
            "categoryCounts": dict(sorted(category_counts.items())),
            "subjectRequirementCounts": dict(sorted(subject_counts.items())),
            "unmatchedGroups": unmatched_groups,
            "notes": [
                "本文件由 scripts/import_shanghai_regular_plan_reference.py 从用户提供 Excel 生成。",
                "院校专业组投档线匹配仍以 data/shanghai/all-admissions.json 中上海考试院官方数据为准。",
                "本文件仅补充组内专业、2026 计划数、学费和 2025 专业录取参考，不用于覆盖官方投档线。",
                "未能在官方 2026 普通批投档线中匹配的专业组保留在 meta.unmatchedGroups。",
            ],
        },
        "groups": groups,
    }


def main() -> None:
    input_path = Path(__import__("sys").argv[1]) if len(__import__("sys").argv) > 1 else DEFAULT_INPUT
    dataset = build_dataset(input_path)
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT_PATH.write_text(json.dumps(dataset, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    meta = dataset["meta"]
    print(
        "wrote {path}: {records} major rows, {groups} groups, {matched} matched official groups, {unmatched} unmatched".format(
            path=OUTPUT_PATH,
            records=meta["recordCount"],
            groups=meta["groupCount"],
            matched=meta["matchedOfficialGroupCount"],
            unmatched=meta["unmatchedGroupCount"],
        )
    )


if __name__ == "__main__":
    main()
